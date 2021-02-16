#! python3
# Automates weekly FedEx shipping reports

import time, os, smtplib, ssl
import openpyxl as xl
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import tkinter as tk
from tkinter import *
from tkinter import messagebox
from tkinter import simpledialog
from pynput.keyboard import Key, Controller
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email import encoders
from email.mime.base import MIMEBase
keyboard = Controller()


#Get FedEx login details from the user with prompt

# Create GUI
root = tk.Tk()
root.eval('tk::PlaceWindow . center') # positions in centre of screen
root.withdraw()
USER_INP = simpledialog.askstring(title="Input", prompt="Fedex Username:") # Gets username
PW_INP = simpledialog.askstring(title="Input", prompt="Fedex Password:", show='*') # Gets password
EMAILPW_INP = simpledialog.askstring(title="Input", prompt="Gmail Password:", show='*') # Gets Gmail PW for later


# Hide automation from Chromium, open in maximized windowp
option = webdriver.ChromeOptions()
option.add_argument('--disable-blink-features=AutomationControlled')
option.add_argument("--start-maximized")

# Set driver varibale
driver = webdriver.Chrome(executable_path='chromedriver.exe', options=option)

# open our webpage
driver.get('https://www.fedex.com/en-us/home.html')
time.sleep(3)

# Select + click our login button
login = driver.find_element_by_xpath('/html/body/div[1]/header/div/div/nav/div/div/div/div[1]/a/span').click()
time.sleep(3)

# Enter username and password
username = driver.find_element_by_xpath('//*[@id="NavLoginUserId"]')
username.send_keys(USER_INP)
username.send_keys(Keys.TAB, PW_INP, Keys.TAB, Keys.TAB, Keys.ENTER)
time.sleep(3)

# Acess Dropdown menu
xShipping = '/html/body/div[1]/header/div/div/nav/div/ul/div[1]/li/a/span'
shipping = driver.find_element_by_xpath(xShipping)
shipping.click()

# CLick "Create a shipment"
xCreate = '/html/body/div[1]/header/div/div/nav/div/ul/div[1]/li/div/div[1]/div/a'
create = driver.find_element_by_xpath(xCreate)
create.click()

# CLick Reports
xReport = '//*[@id="menubar.nav.menu5_div"]'
report = driver.find_element_by_xpath(xReport)
report.click()
time.sleep(3)

# Check the right boxes, how to loop by xpath??
shipdate = driver.find_element_by_xpath('//*[@id="fields"]/tbody/tr[3]/td[1]/input[1]').click()
track = driver.find_element_by_xpath('//*[@id="fields"]/tbody/tr[3]/td[1]/input[2]').click()
servtype = driver.find_element_by_xpath('//*[@id="fields"]/tbody/tr[3]/td[1]/input[3]').click()
paytype = driver.find_element_by_xpath('//*[@id="fields"]/tbody/tr[3]/td[1]/input[4]').click()
billtran = driver.find_element_by_xpath('//*[@id="fields"]/tbody/tr[3]/td[1]/input[5]').click()
billdtf = driver.find_element_by_xpath('//*[@id="fields"]/tbody/tr[3]/td[1]/input[6]').click()
send = driver.find_element_by_xpath('//*[@id="fields"]/tbody/tr[3]/td[1]/input[7]').click()
rec = driver.find_element_by_xpath('//*[@id="fields"]/tbody/tr[3]/td[1]/input[8]').click()
ref = driver.find_element_by_xpath('//*[@id="fields"]/tbody/tr[3]/td[1]/input[10]').click()

# Alert user instructions
messagebox.showinfo('Automation Paused 20 seconds', 'Please click OK, choose the start and end dates, then wait.\n\nScript will terminate if there are no shipments to report.')
time.sleep(15)
submit = driver.find_element_by_xpath('//*[@id="content"]/table[3]/tbody/tr[4]/td[4]/p/input')
submit.click()

# Download report in csv
time.sleep(3)
download = driver.find_element_by_xpath('//*[@id="content"]/table[3]/tbody/tr[4]/td[2]/p/input[2]').click()

# Launch our downloaded csv file
time.sleep(3)
os.startfile("C:\\File\\Path\\Here\\download.csv")
time.sleep(3)

# Enable access
keyboard.press(Key.left)
keyboard.press(Key.enter)
time.sleep(3)

#Enable Saving
keyboard.press(Key.f12)
time.sleep(1)
keyboard.press(Key.left)
time.sleep(1)
keyboard.press(Key.enter)
time.sleep(1)

# Save as .xlsx file
keyboard.type('fedexDownload')
keyboard.press(Key.tab)
time.sleep(1)
keyboard.press(Key.down)
time.sleep(1)

# Ensure saved as the file format at the top of the list in the window
for i in range(0, 6):
    keyboard.press(Key.up)
    time.sleep(1)
keyboard.press(Key.enter)
time.sleep(1)
keyboard.press(Key.enter)

# Close file
keyboard.press(Key.alt)
keyboard.press(Key.f4)
keyboard.release(Key.alt)
keyboard.release(Key.f4)
time.sleep(2)

# delete csv file
os.remove("C:\\File\\Path\\Here\\download.csv")

# Open the new xlsx file
new_xlsx = 'C:\\File\\Path\\Here\\fedexDownload.xlsx'
wb = xl.load_workbook(new_xlsx)
ws = wb.worksheets[0]

# Delete unnecessary rows
ws.delete_cols(4)
ws.delete_cols(5)
ws.delete_cols(6,6)
ws.delete_cols(7,9)

# Reorder columns to match report
ws.move_range("A1:F50", rows=0, cols=14)
ws['A1'] = ws.cell(1, 15).value
ws['A2'] = ws.cell(2, 15).value
ws['A3'] = ws.cell(3, 15).value
ws['A4'] = ws.cell(4, 15).value
ws['A5'] = ws.cell(5, 15).value
ws['A6'] = ws.cell(6, 15).value
ws['A7'] = ws.cell(7, 15).value
ws['A8'] = ws.cell(8, 15).value
ws['A9'] = ws.cell(9, 15).value
ws['A10'] = ws.cell(10, 15).value
ws['B1'] = ws.cell(1, 16).value
ws['B2'] = ws.cell(2, 16).value
ws['B3'] = ws.cell(3, 16).value
ws['C1'] = ws.cell(1, 19).value
ws['C2'] = ws.cell(2, 19).value
ws['C3'] = ws.cell(3, 19).value
ws['D1'] = ws.cell(1, 18).value
ws['D2'] = ws.cell(2, 18).value
ws['D3'] = ws.cell(3, 18).value
ws['E1'] = ws.cell(1, 20).value
ws['E2'] = ws.cell(2, 20).value
ws['E3'] = ws.cell(3, 20).value
ws['G1'] = ws.cell(1, 17).value
ws['G2'] = ws.cell(2, 17).value
ws['G3'] = ws.cell(3, 17).value
ws['J1'] = ws.cell(1, 3).value
ws['J2'] = ws.cell(2, 4).value
ws['J3'] = ws.cell(3, 4).value
ws.delete_cols(15,6)
ws.delete_rows(1, 1)
wb.save(str(new_xlsx))
wb.close()

# Open source file
src_filename = 'C:\\File\\Path\\Here\\fedexDownload.xlsx'
wb1 = xl.load_workbook(src_filename)
ws1 = wb1.worksheets[0]

# Open destination file
des_filename = 'C:\\File\\Path\\Here\\realShipReport.xlsx'
wb2 = xl.load_workbook(des_filename)
ws2 = wb2.active

# Calculate total number of rows and columns in source excel file
mr = ws1.max_row
mc = ws1.max_column

# copying the cell values from source excel file to destination excel file
for i in range(1, mr + 1):
    for j in range(1, mc + 1):
        # reading cell value from source excel file
        c = ws1.cell(row=i, column=j)

        # writing the read value to destination excel file
        ws2.cell(row=691 + i, column=j).value = c.value

# saving the destination excel file
wb2.save(str(des_filename))
wb2.close()

# delete source file
os.remove("C:\\File\\Path\\Here\\fedexDownload.xlsx")

# EMAIL the file to work computer
subject = "Python Automated Ship Report"
body = "This is an email with attachment sent from Python"
sender_email = "your.email@gmail.com"
receiver_email = "receiver.email@gmail.com"
password = EMAILPW_INP 

# Create a multipart message and set headers
message = MIMEMultipart()
message["From"] = sender_email
message["To"] = receiver_email
message["Subject"] = subject
message["Bcc"] = receiver_email  # Recommended for mass emails

# Add body to email
message.attach(MIMEText(body, "plain"))

filename = "C:\\File\\Path\\Here\\realShipReport.xlsx"

# Open file in binary mode
with open(filename, "rb") as attachment:
    # Add file as application/octet-stream
    # Email client can usually download this automatically as attachment
    part = MIMEBase("application", "octet-stream")
    part.set_payload(attachment.read())

# Encode file in ASCII characters to send by email
encoders.encode_base64(part)

# Add header as key/value pair to attachment part
part.add_header(
    "Content-Disposition",
    f"attachment; filename= {filename}",
)

# Add attachment to message and convert message to string
message.attach(part)
text = message.as_string()

# Log in to server using secure context and send email
context = ssl.create_default_context()
with smtplib.SMTP_SSL("smtp.gmail.com", 465, context=context) as server:
    server.login(sender_email, password)
    server.sendmail(sender_email, receiver_email, text)

messagebox.showinfo('Script Finished', 'File sent')






